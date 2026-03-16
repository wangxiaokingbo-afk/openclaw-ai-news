#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "Skills 使用指南"

# 数据
headers = ["序号", "Skill", "图标", "用途", "难度", "配置要求", "典型命令"]

data = [
    [1, "openclaw-tavily-search", "🌐", "Tavily 网络搜索", "⭐", "TAVILY_API_KEY", 'python3 tavily_search.py --query "..."'],
    [2, "find-skills", "🔍", "发现新 skills", "⭐", "无", "npx skills find <query>"],
    [3, "skill-vetting", "🛡️", "安全审查 skills", "⭐⭐", "Python 3", "python3 scan.py ."],
    [4, "self-improving", "🧠", "自我反思 + 记忆", "⭐⭐⭐", "初始化目录", "自动触发"],
    [5, "summarize", "🧾", "总结 URL/文件/视频", "⭐", "各模型 API Key", 'summarize "url" --length short'],
    [6, "using-superpowers", "⚡", "强制使用 skills", "⭐⭐", "无", "自动触发（1% 可能就要用）"],
    [7, "github", "🐙", "GitHub CLI 交互", "⭐⭐", "gh CLI + 认证", "gh pr checks <num>"],
    [8, "agent-browser", "🌍", "浏览器自动化", "⭐⭐⭐", "Node.js + npm", "agent-browser open <url>"],
]

# 样式
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 设置列宽
column_widths = [6, 28, 6, 18, 8, 18, 35]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = width

# 设置行高
ws.row_dimensions[1].height = 25

# 写入表头
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 写入数据
for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border
    ws.row_dimensions[row_idx].height = 30

# 添加第二个表格 - 使用场景
ws2 = wb.create_sheet(title="使用场景")

scenario_headers = ["需求", "推荐 Skill", "命令示例"]
scenario_data = [
    ["搜索最新资料", "🌐 tavily-search", '--query "AI 最佳实践" --max-results 5'],
    ["找新技能扩展", "🔍 find-skills", "npx skills find react testing"],
    ["安装前安全检查", "🛡️ skill-vetting", "curl + unzip + scan.py ."],
    ["记住用户偏好", "🧠 self-improving", "自动记录纠正 → 存入 memory.md"],
    ["快速读长文章", "🧾 summarize", 'summarize "https://..." --length short'],
    ["确保不漏技能", "⚡ using-superpowers", "自动触发（1% 可能就要用）"],
    ["查 PR/CI 状态", "🐙 github", "gh pr checks 55 --repo owner/repo"],
    ["网页自动化", "🌍 agent-browser", "open → snapshot -i → click @e1"],
]

# 表头样式
for col, header in enumerate(scenario_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 数据
for row_idx, row_data in enumerate(scenario_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border
    ws2.row_dimensions[row_idx].height = 28

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 40

# 添加第三个表格 - 注意事项
ws3 = wb.create_sheet(title="注意事项")

note_headers = ["Skill", "注意事项"]
note_data = [
    ["🌐 tavily-search", "需要 API Key，保持 max-results ≤ 5"],
    ["🛡️ skill-vetting", "安装前必用，发现 eval/base64/未知域名→拒绝"],
    ["🧠 self-improving", "记忆分 HOT/WARM/COLD 三层，定期整理"],
    ["🧾 summarize", "支持 PDF/YouTube，需设置对应模型 API Key"],
    ["⚡ using-superpowers", "不可跳过，即使觉得很简单也要检查"],
    ["🐙 github", "需先 gh auth login 认证"],
    ["🌍 agent-browser", "refs 导航后变化，每次重新 snapshot"],
]

for col, header in enumerate(note_headers, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for row_idx, row_data in enumerate(note_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border
    ws3.row_dimensions[row_idx].height = 35

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 50

# 保存文件
output_path = "/Users/ssd/.openclaw/workspace/docs/Skills 使用指南.xlsx"
wb.save(output_path)
print(f"✅ Excel 文件已创建：{output_path}")
