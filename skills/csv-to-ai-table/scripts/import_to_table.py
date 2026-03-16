#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV/JSON 文件导入到钉钉 AI 表格
通用脚本 - 可被其他 agent 调用
"""

import argparse
import json
import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path


# ==================== 配置 ====================

WORKSPACE = os.environ.get("OPENCLAW_WORKSPACE", str(Path.home() / ".openclaw" / "workspace"))
TABLE_MD = Path(WORKSPACE) / "TABLE.md"
MAX_FILE_SIZE_CSV = 50 * 1024 * 1024  # 50MB
MAX_FILE_SIZE_JSON = 10 * 1024 * 1024  # 10MB


# ==================== 工具函数 ====================

def safe_path(file_path):
    """安全路径解析，防止目录遍历攻击"""
    path = Path(file_path).resolve()
    workspace = Path(WORKSPACE).resolve()
    
    try:
        path.relative_to(workspace)
        return path
    except ValueError:
        # 允许绝对路径但记录警告
        if path.exists():
            print(f"⚠️  警告：文件在工作区外：{path}")
            return path
        raise FileNotFoundError(f"文件不存在：{file_path}")


def get_root_uuid():
    """获取根节点 UUID"""
    if not TABLE_MD.exists():
        print("❌ TABLE.md 不存在，请先获取根节点")
        print("运行：mcporter call dingtalk-ai-table get_root_node_of_my_document --output json")
        sys.exit(1)
    
    content = TABLE_MD.read_text(encoding='utf-8')
    match = re.search(r'rootDentryUuid[:\s]+([a-zA-Z0-9]+)', content)
    
    if not match:
        print("❌ TABLE.md 中未找到 rootDentryUuid")
        sys.exit(1)
    
    return match.group(1)


def detect_field_type(values, field_name):
    """自动推断字段类型"""
    # 检查字段名提示
    name_lower = field_name.lower()
    if any(k in name_lower for k in ['日期', '时间', 'date', 'time']):
        return 'date'
    if any(k in name_lower for k in ['负责人', '用户', '人员', 'user', 'owner']):
        return 'user'
    
    # 检查值类型
    non_empty = [v for v in values if v is not None and v != '']
    
    if not non_empty:
        return 'text'
    
    # 检查是否都是数字
    all_numeric = all(
        isinstance(v, (int, float)) or 
        (isinstance(v, str) and re.match(r'^-?\d+(\.\d+)?$', v.strip()))
        for v in non_empty
    )
    
    if all_numeric:
        return 'number'
    
    # 检查是否有状态标记
    status_markers = ['✅', '❌', '⚠️', '✓', '×', '待', '已', '未']
    has_status = any(
        isinstance(v, str) and any(m in v for m in status_markers)
        for v in non_empty
    )
    
    if has_status:
        return 'singleSelect'
    
    return 'text'


def read_csv(file_path):
    """读取 CSV 文件"""
    import csv
    
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        return list(reader)


def read_excel(file_path):
    """读取 Excel 文件"""
    from openpyxl import load_workbook
    
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active
    
    # 获取表头
    headers = [cell.value for cell in ws[1]]
    
    # 获取数据行
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                record[headers[i]] = value
        if any(v is not None for v in record.values()):
            records.append(record)
    
    wb.close()
    return records


def read_json(file_path):
    """读取 JSON 文件"""
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 支持 {data: [...]} 格式
    if isinstance(data, dict) and 'data' in data:
        return data['data']
    
    return data


def infer_fields(records):
    """从记录中推断字段配置"""
    if not records:
        return []
    
    fields = []
    field_names = list(records[0].keys())
    
    for field_name in field_names:
        # 跳过空字段名
        if not field_name:
            continue
        
        values = [r.get(field_name) for r in records]
        field_type = detect_field_type(values, field_name)
        
        fields.append({
            "name": field_name,
            "type": field_type
        })
    
    return fields


def create_table(table_name, root_uuid, sheet_name="数据表", fields=None):
    """创建 AI 表格"""
    print(f"📋 创建表格：{table_name}")
    
    # 1. 创建表格
    cmd = [
        "mcporter", "call", "dingtalk-ai-table", "create_base_app",
        f"filename={table_name}",
        f"target={root_uuid}",
        "--output", "json"
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    
    if result.returncode != 0:
        print(f"❌ 创建表格失败：{result.stderr}")
        return None
    
    table_info = json.loads(result.stdout)
    dentry_uuid = table_info.get('body', {}).get('info', {}).get('uuid')
    
    if not dentry_uuid:
        print(f"❌ 未获取到表格 UUID: {result.stdout}")
        return None
    
    print(f"✅ 表格创建成功，UUID: {dentry_uuid}")
    
    # 2. 创建数据表和字段
    if fields:
        print(f"📊 创建数据表：{sheet_name}，字段数：{len(fields)}")
        
        args = {
            "dentryUuid": dentry_uuid,
            "name": sheet_name,
            "fields": fields
        }
        
        cmd = [
            "mcporter", "call", "dingtalk-ai-table", "add_base_table",
            "--args", json.dumps(args, ensure_ascii=False),
            "--output", "json"
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        
        if result.returncode != 0:
            print(f"⚠️  创建数据表失败：{result.stderr}")
        else:
            print(f"✅ 数据表创建成功")
    
    return dentry_uuid


def import_records(dentry_uuid, sheet_name, records, batch_size=100):
    """批量导入记录"""
    print(f"📥 导入记录到 {sheet_name}，共 {len(records)} 条")
    
    total = len(records)
    imported = 0
    
    for i in range(0, total, batch_size):
        batch = records[i:i+batch_size]
        
        # 构建记录参数
        mcporter_records = []
        for record in batch:
            fields = {}
            for key, value in record.items():
                # 清理字段值
                if value == '' or value is None:
                    continue
                fields[key] = value
            mcporter_records.append({"fields": fields})
        
        args = {
            "dentryUuid": dentry_uuid,
            "sheetIdOrName": sheet_name,
            "records": mcporter_records
        }
        
        cmd = [
            "mcporter", "call", "dingtalk-ai-table", "add_base_record",
            "--args", json.dumps(args, ensure_ascii=False),
            "--output", "json"
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        
        if result.returncode != 0:
            print(f"⚠️  批次 {i//batch_size + 1} 导入失败：{result.stderr}")
        else:
            imported += len(batch)
            print(f"✅ 批次 {i//batch_size + 1}: 导入 {len(batch)} 条记录")
    
    return imported


# ==================== 主函数 ====================

def import_file_to_table(file_path, table_name, sheet_name="数据表", fields=None, dry_run=False):
    """主导入函数"""
    print("="*60)
    print("🚀 CSV/JSON 文件导入到钉钉 AI 表格")
    print("="*60)
    
    # 1. 验证文件
    file_path = safe_path(file_path)
    
    if not file_path.exists():
        print(f"❌ 文件不存在：{file_path}")
        return {"success": False, "error": "文件不存在"}
    
    # 检查文件大小
    file_size = file_path.stat().st_size
    if file_path.suffix.lower() == '.csv' and file_size > MAX_FILE_SIZE_CSV:
        print(f"❌ CSV 文件过大（最大 50MB）")
        return {"success": False, "error": "文件过大"}
    
    if file_path.suffix.lower() == '.json' and file_size > MAX_FILE_SIZE_JSON:
        print(f"❌ JSON 文件过大（最大 10MB）")
        return {"success": False, "error": "文件过大"}
    
    print(f"📂 文件：{file_path}")
    print(f"📊 大小：{file_size / 1024:.1f} KB")
    
    # 2. 读取数据
    print("\n📖 读取数据...")
    
    if file_path.suffix.lower() in ['.csv', '.txt']:
        records = read_csv(file_path)
    elif file_path.suffix.lower() in ['.xlsx', '.xls']:
        records = read_excel(file_path)
    elif file_path.suffix.lower() == '.json':
        records = read_json(file_path)
    else:
        print(f"❌ 不支持的文件类型：{file_path.suffix}")
        return {"success": False, "error": "不支持的文件类型"}
    
    print(f"✅ 读取 {len(records)} 条记录")
    
    if not records:
        print("⚠️  数据为空")
        return {"success": False, "error": "数据为空"}
    
    # 3. 推断字段
    if not fields:
        print("\n🔍 推断字段类型...")
        fields = infer_fields(records)
        
        for f in fields:
            print(f"  - {f['name']}: {f['type']}")
    
    # 4. 预览模式
    if dry_run:
        print("\n👁️  预览模式 - 不执行实际导入")
        print(f"表格名称：{table_name}")
        print(f"数据表：{sheet_name}")
        print(f"字段数：{len(fields)}")
        print(f"记录数：{len(records)}")
        return {"success": True, "dry_run": True}
    
    # 5. 获取根节点
    print("\n🔑 获取根节点...")
    root_uuid = get_root_uuid()
    print(f"✅ 根节点 UUID: {root_uuid}")
    
    # 6. 创建表格
    print("\n📋 创建表格...")
    dentry_uuid = create_table(table_name, root_uuid, sheet_name, fields)
    
    if not dentry_uuid:
        return {"success": False, "error": "创建表格失败"}
    
    # 7. 导入记录
    print("\n📥 导入记录...")
    imported = import_records(dentry_uuid, sheet_name, records)
    
    # 8. 结果
    print("\n" + "="*60)
    print("✅ 导入完成！")
    print("="*60)
    print(f"表格名称：{table_name}")
    print(f"数据表：{sheet_name}")
    print(f"导入记录：{imported}/{len(records)}")
    print(f"表格 UUID: {dentry_uuid}")
    
    return {
        "success": True,
        "table_name": table_name,
        "sheet_name": sheet_name,
        "dentry_uuid": dentry_uuid,
        "record_count": imported,
        "total_records": len(records)
    }


def main():
    parser = argparse.ArgumentParser(
        description="CSV/JSON 文件导入到钉钉 AI 表格",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python3 import_to_table.py --file data.csv --name "我的表格"
  python3 import_to_table.py --file data.json --name "销售数据" --sheet "销售记录"
  python3 import_to_table.py --file data.csv --name "测试" --dry-run
        """
    )
    
    parser.add_argument("--file", required=True, help="CSV 或 JSON 文件路径")
    parser.add_argument("--name", required=True, help="AI 表格名称")
    parser.add_argument("--sheet", default="数据表", help="数据表名称（默认：数据表）")
    parser.add_argument("--fields", help="字段配置 JSON（可选，自动推断）")
    parser.add_argument("--batch-size", type=int, default=100, help="批次大小（默认：100）")
    parser.add_argument("--dry-run", action="store_true", help="预览模式，不实际导入")
    
    args = parser.parse_args()
    
    # 解析字段配置
    fields = None
    if args.fields:
        fields = json.loads(args.fields)
    
    # 执行导入
    result = import_file_to_table(
        file_path=args.file,
        table_name=args.name,
        sheet_name=args.sheet,
        fields=fields,
        dry_run=args.dry_run
    )
    
    # 返回状态码
    sys.exit(0 if result.get("success") else 1)


if __name__ == "__main__":
    main()
